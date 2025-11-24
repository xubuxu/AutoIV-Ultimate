"""
AutoIV-Ultimate Reporting Module - Excel Builder

Creates comprehensive Excel reports with statistics, champion cells, and yield data.
"""
import logging
from pathlib import Path
from typing import Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelBuilder:
    """Generates Excel reports with formatted tables."""
    
    def __init__(self, output_dir: str = "output"):
        """
        Initialize ExcelBuilder.
        
        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info("ExcelBuilder initialized")
    
    def create_report(self, stats_df: pd.DataFrame, champion_df: pd.DataFrame,
                     top_cells_df: pd.DataFrame, yield_df: pd.DataFrame,
                     clean_df: pd.DataFrame, filename: str = "IV_Analysis_Report.xlsx") -> str:
        """
        Create comprehensive Excel report.
        
        Args:
            stats_df: Statistics dataframe
            champion_df: Champion cells dataframe
            top_cells_df: Top 10 cells dataframe
            yield_df: Yield distribution dataframe
            clean_df: Cleaned raw data
            filename: Output filename
            
        Returns:
            Path to saved Excel file
        """
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets
            self._add_statistics_sheet(wb, stats_df)
            self._add_champion_sheet(wb, champion_df)
            self._add_top_cells_sheet(wb, top_cells_df)
            self._add_yield_sheet(wb, yield_df)
            self._add_raw_data_sheet(wb, clean_df)
            
            # Save
            output_path = self.output_dir / filename
            wb.save(output_path)
            
            logger.info(f"Excel report saved: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Excel report generation failed: {e}")
            return ""
    
    def _add_statistics_sheet(self, wb: Workbook, stats_df: pd.DataFrame):
        """Add statistics summary sheet."""
        ws = wb.create_sheet("Statistics Summary")
        
        # Add title
        ws['A1'] = "Batch Statistics Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add dataframe
        for r_idx, row in enumerate(dataframe_to_rows(stats_df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatting
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_champion_sheet(self, wb: Workbook, champion_df: pd.DataFrame):
        """Add champion cells sheet."""
        ws = wb.create_sheet("Champion Cells")
        
        ws['A1'] = "Champion Cells (Best per Batch)"
        ws['A1'].font = Font(size=14, bold=True)
        
        for r_idx, row in enumerate(dataframe_to_rows(champion_df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
    
    def _add_top_cells_sheet(self, wb: Workbook, top_cells_df: pd.DataFrame):
        """Add top 10 cells sheet."""
        ws = wb.create_sheet("Top 10 Cells")
        
        ws['A1'] = "Top 10 Cells Overall"
        ws['A1'].font = Font(size=14, bold=True)
        
        for r_idx, row in enumerate(dataframe_to_rows(top_cells_df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    def _add_yield_sheet(self, wb: Workbook, yield_df: pd.DataFrame):
        """Add yield distribution sheet."""
        ws = wb.create_sheet("Yield Distribution")
        
        ws['A1'] = "Yield Distribution by Efficiency Bins"
        ws['A1'].font = Font(size=14, bold=True)
        
        for r_idx, row in enumerate(dataframe_to_rows(yield_df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    def _add_raw_data_sheet(self, wb: Workbook, clean_df: pd.DataFrame):
        """Add cleaned raw data sheet."""
        ws = wb.create_sheet("Cleaned Data")
        
        ws['A1'] = "Cleaned Raw Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Limit to first 1000 rows to avoid Excel file size issues
        df_limited = clean_df.head(1000)
        
        for r_idx, row in enumerate(dataframe_to_rows(df_limited, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True)
